#!/usr/bin/env python3
"""
営業ログから自動でLP生成
画像を雰囲気に合わせて自動割り当て
"""

import csv
import os
import random
from datetime import datetime
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: openpyxlがインストールされていません。pip install openpyxlを実行してください。")


# 画像プール（雰囲気別）
IMAGE_POOLS = {
    'elegant': list(range(1, 9)),      # image01-08: エレガント・上品
    'natural': list(range(9, 17)),     # image09-16: ナチュラル・癒し
    'modern': list(range(17, 26)),     # image17-25: モダン・洗練
}

THERAPIST_IMAGES = list(range(26, 34))  # image26-33: 施術者画像


def detect_atmosphere(brand_name, reference_url=''):
    """
    ブランド名やURLから雰囲気を推測
    """
    brand_name = str(brand_name) if brand_name else ''
    reference_url = str(reference_url) if reference_url else ''
    text = (brand_name + ' ' + reference_url).lower()

    # キーワードマッチング
    if any(word in text for word in ['モダン', 'modern', 'studio', 'スタジオ', 'シンプル']):
        return 'modern'
    elif any(word in text for word in ['ナチュラル', 'natural', '森', '緑', 'グリーン', '癒し']):
        return 'natural'
    else:
        return 'elegant'  # デフォルト


def assign_images(atmosphere):
    """
    雰囲気に応じて画像を割り当て
    """
    pool = IMAGE_POOLS.get(atmosphere, IMAGE_POOLS['elegant'])
    main_image_num = random.choice(pool)
    therapist_image_num = random.choice(THERAPIST_IMAGES)

    return (
        f'image{main_image_num:02d}.jpg',
        f'image{therapist_image_num:02d}.jpg'
    )


def _image_number(filename):
    if not filename.startswith('image') or not filename.endswith('.jpg'):
        return None
    num_part = filename[5:-4]
    if not num_part.isdigit():
        return None
    return int(num_part)


def get_even_background_cycle(template):
    """
    output/<template>/images から背景画像を最大3枚選び、均等配分用のリストを返す
    """
    images_dir = os.path.join('output', template, 'images')
    if not os.path.isdir(images_dir):
        return []
    try:
        available = set(os.listdir(images_dir))
    except Exception:
        return []

    preferred_nums = [20, 21, 22]
    preferred = []
    for n in preferred_nums:
        fname = f'image{n:02d}.jpg'
        if fname in available:
            preferred.append(fname)

    backgrounds = list(preferred)
    if len(backgrounds) < 3:
        others = []
        for fname in sorted(available):
            if fname in backgrounds:
                continue
            num = _image_number(fname)
            if num is None:
                continue
            if 1 <= num <= 25:
                others.append(fname)
        for fname in others:
            if len(backgrounds) >= 3:
                break
            backgrounds.append(fname)

    return backgrounds


def read_excel(file_path):
    """
    Excelファイルを読み込む
    """
    if not EXCEL_AVAILABLE:
        print("エラー: openpyxlがインストールされていません。")
        print("以下のコマンドを実行してください:")
        print("  pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # ヘッダー行を取得
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # データ行を取得
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 最初の列が空でない行のみ
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            rows.append(row_dict)

    return rows


def normalize_column_names(row):
    """
    列名を正規化（全角・半角、大文字・小文字の違いを吸収）
    """
    normalized = {}

    # 店名・ブランド名の検出
    for key in row.keys():
        if not key:
            continue

        key_str = str(key).strip()

        # 完全一致チェック
        if key_str in ['店名', 'ブランド名', 'brand_name', 'name']:
            if row[key] and str(row[key]).strip():
                val = str(row[key]).strip()
                # Excel may contain formulas like =IFERROR(...). If so, try to
                # extract a quoted fallback value; otherwise treat as empty.
                if val.startswith('='):
                    import re
                    m = re.search(r'IFERROR\([^,]*,\s*"([^"]*)"\)', val, re.IGNORECASE)
                    if m and m.group(1):
                        val = m.group(1).strip()
                    else:
                        # remove common formula tokens and non-printables
                        val = re.sub(r'[=\"\(\)\[\]]', '', val).strip()
                        # if remaining text looks like a formula or is empty, blank it
                        if not val or any(tok in val.upper() for tok in ['IFERROR', 'VLOOKUP', 'INDEX', 'MATCH']):
                            val = ''

                normalized['brand_name'] = val

        # URL検出（複数のURLカラムがある場合、最初に見つかったものを使用）
        elif 'url' in key_str.lower() and 'reference_url' not in normalized:
            if row[key] and str(row[key]).strip():
                url_val = str(row[key]).strip()
                if url_val.startswith('http'):
                    normalized['reference_url'] = url_val

        # ID検出
        elif key_str.lower() == 'id':
            if row[key]:
                normalized['id'] = str(row[key]).strip()

    # IDがない場合は自動生成
    if 'id' not in normalized:
        if 'brand_name' in normalized:
            # ブランド名から数字部分を抽出、なければランダム
            import re
            numbers = re.findall(r'\d+', str(normalized.get('brand_name', '')))
            if numbers:
                normalized['id'] = f"{int(numbers[0]):04d}"
            else:
                # ブランド名のハッシュから生成
                hash_val = hash(normalized.get('brand_name', '')) % 10000
                normalized['id'] = f"{hash_val:04d}"
        else:
            normalized['id'] = f"{random.randint(1, 9999):04d}"

    # URLがない場合はデフォルト
    if 'reference_url' not in normalized:
        normalized['reference_url'] = 'https://example.com'

    return normalized


def read_template(template_name):
    """テンプレートファイルを読み込む"""
    template_path = os.path.join('templates', f'variant{template_name}.html')
    with open(template_path, 'r', encoding='utf-8') as f:
        return f.read()


def replace_placeholders(template, brand_name, image_url, reference_url, year, therapist_image_url):
    """プレースホルダーを実際の値に置換"""
    result = template.replace('{{BRAND_NAME}}', brand_name)
    result = result.replace('{{IMAGE_URL}}', image_url)
    result = result.replace('{{REFERENCE_URL}}', reference_url)
    result = result.replace('{{YEAR}}', str(year))
    result = result.replace('{{THERAPIST_IMAGE_URL}}', therapist_image_url)
    return result


def generate_html_from_row(row_data, year, template='A', output_template=None, forced_main_image=None):
    """
    1行分のデータからHTMLを生成
    """
    # 列名を正規化
    normalized = normalize_column_names(row_data)

    if 'brand_name' not in normalized or not normalized['brand_name']:
        print(f"警告: ブランド名が見つかりません: {row_data}")
        return None

    id_val = normalized.get('id', f"{random.randint(1, 9999):04d}")
    # sanitize brand_name to avoid embedding raw Excel formulas like =IFERROR(...)
    import re
    def sanitize_text(s):
        if not s:
            return ''
        s = str(s).strip()
        # if starts with '=' try to capture quoted fallback text
        if s.startswith('='):
            quotes = re.findall(r'"([^\"]+)"', s)
            if quotes:
                # pick last quoted value (often the fallback in IFERROR)
                candidate = quotes[-1].strip()
                # ignore placeholder tokens
                if candidate and not any(tok in candidate.upper() for tok in ['COMPUTED_VALUE', '__XLUDF', '__XL']):
                    return candidate
            # otherwise remove common formula tokens
            cleaned = re.sub(r'[=\"\(\)]', '', s).strip()
            if 'IFERROR' in cleaned.upper() or 'VLOOKUP' in cleaned.upper():
                return ''
            return cleaned
        # remove obvious formula remnants
        if 'IFERROR' in s.upper() or s.startswith('#'):
            return ''
        return s

    brand_name = sanitize_text(normalized.get('brand_name', ''))
    if not brand_name:
        print(f"警告: ブランド名が見つかりません (プレースホルダを使用します): {row_data}")
        brand_name = 'サロン'

    reference_url = str(normalized.get('reference_url', 'https://example.com'))

    # 雰囲気を推測して画像を割り当て
    atmosphere = detect_atmosphere(brand_name, reference_url)
    main_image, therapist_image = assign_images(atmosphere)
    if forced_main_image:
        main_image = forced_main_image

    output_template = (output_template or template).strip().upper()

    # Ensure selected images actually exist in output/<template>/images/
    images_dir = os.path.join('output', output_template, 'images')
    available = set()
    if os.path.isdir(images_dir):
        try:
            available = set(os.listdir(images_dir))
        except Exception:
            available = set()

    # Helper to prefer images from a numeric pool present in available files
    def choose_from_pool(pool_numbers, prefer=None):
        candidates = []
        for n in pool_numbers:
            fname = f'image{n:02d}.jpg'
            if fname in available:
                candidates.append(fname)
        if candidates:
            # if prefer is available, use it
            if prefer and prefer in candidates:
                return prefer
            return random.choice(candidates)
        return None

    # Validate main image: if missing, try to pick from IMAGE_POOLS for this atmosphere
    if main_image not in available:
        pool_nums = IMAGE_POOLS.get(atmosphere, IMAGE_POOLS['elegant'])
        chosen = choose_from_pool(pool_nums, prefer=main_image)
        if chosen:
            main_image = chosen
        else:
            # fallback to any available image
            if available:
                main_image = sorted(list(available))[0]

    # Validate therapist image: prefer therapist pool (26-33), else pick any available
    if therapist_image not in available:
        chosen = choose_from_pool(THERAPIST_IMAGES, prefer=therapist_image)
        if chosen:
            therapist_image = chosen
        else:
            # if no therapist images available, fallback to main_image
            if main_image in available:
                therapist_image = main_image
            elif available:
                therapist_image = sorted(list(available))[0]

    # 出力先ディレクトリ（テンプレート別）を用意
    output_dir = os.path.join('output', output_template)
    os.makedirs(output_dir, exist_ok=True)

    # 画像は output/<template>/images に置かれる想定なので、HTMLから見た相対パスを設定
    image_url = f'images/{main_image}'
    therapist_image_url = f'images/{therapist_image}'

    # テンプレート読み込み
    template_content = read_template(template)

    # プレースホルダー置換
    html_content = replace_placeholders(
        template_content,
        brand_name,
        image_url,
        reference_url,
        year,
        therapist_image_url
    )

    # 出力ファイル名 (例: 08000A.html)
    output_filename = f'{id_val}{output_template}.html'
    output_path = os.path.join(output_dir, output_filename)

    # ファイル書き出し
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f'[OK] {os.path.join(output_dir, output_filename)} を生成しました ({atmosphere}の雰囲気: {main_image})')

    return {
        'id': id_val,
        'brand_name': brand_name,
        'reference_url': reference_url,
        'template': output_template,
        'image': main_image,
        'therapist_image': therapist_image,
        'atmosphere': atmosphere
    }


def main():
    """メイン処理"""
    # 現在の年を取得
    current_year = datetime.now().year

    # outputディレクトリが存在しない場合は作成
    os.makedirs('output', exist_ok=True)

    # コマンドライン引数: --start-id を受け取り、指定があれば全行に連番IDを割り当てる
    import argparse
    parser = argparse.ArgumentParser(description='営業ログから自動でLP生成')
    parser.add_argument('--start-id', type=int, default=None, help='IDを指定した番号から採番開始 (例: 1000)')
    parser.add_argument('--template', '-t', type=str, default='A', help='テンプレート種別 (A/B/C...)')
    parser.add_argument('--output-template', type=str, default=None, help='出力先のテンプレート記号 (例: B)')
    parser.add_argument('--excel', type=str, default=None, help='使用する営業ログExcelファイルのパス')
    args = parser.parse_args()
    start_id = args.start_id
    template = (args.template or 'A').strip().upper()
    output_template = (args.output_template or template).strip().upper()
    start_counter = 0

    # 営業ログファイルを探す
    excel_files = []
    for root, dirs, files in os.walk('input'):
        for file in files:
            if file.endswith(('.xlsx', '.xls')) and '営業' in file:
                excel_files.append(os.path.join(root, file))

    if args.excel:
        if os.path.isfile(args.excel):
            excel_files = [args.excel]
        else:
            print(f"エラー: 指定されたExcelファイルが見つかりません: {args.excel}")
            return

    if not excel_files:
        print("エラー: 営業ログファイルが見つかりません")
        print("input/フォルダに「営業」という文字を含むExcelファイルを配置してください")
        return

    # 最初のファイルを使用
    excel_file = excel_files[0]
    print(f"営業ログを読み込んでいます: {excel_file}")

    # Excelを読み込む
    rows = read_excel(excel_file)

    if not rows:
        print("エラー: データが読み込めませんでした")
        return

    print(f"\n{len(rows)}件のデータを検出しました")
    print("HTMLを生成しています...\n")

    # NOTE: Do not copy images. Images are expected to already exist
    # under the appropriate output folder (e.g. output/A/images/).

    # 背景画像を均等配分（variantAのみ）
    bg_cycle = []
    if template == 'A':
        bg_cycle = get_even_background_cycle(output_template)
        if bg_cycle:
            print(f"背景画像を均等配分: {', '.join(bg_cycle)}")

    # 生成結果を記録
    results = []

    # 各行からHTMLを生成
    for i, row in enumerate(rows, 1):
        print(f"[{i}/{len(rows)}] ", end="")

        # start-id が指定されていれば、全行に連番IDを割り当てる（既存IDを上書き）
        if start_id is not None:
            seq_id = start_id + start_counter
            row['id'] = f"{seq_id:05d}"
            start_counter += 1

        forced_main = None
        if bg_cycle:
            forced_main = bg_cycle[(i - 1) % len(bg_cycle)]

        result = generate_html_from_row(
            row,
            current_year,
            template=template,
            output_template=output_template,
            forced_main_image=forced_main
        )
        if result:
            results.append(result)

    print(f"\n完了！{len(results)}件のHTMLファイルを生成しました")
    print(f"output/ フォルダを確認してください")

    # 生成ログをCSVに出力
    log_file = 'output/generation_log.csv'
    with open(log_file, 'w', encoding='utf-8', newline='') as f:
        if results:
            writer = csv.DictWriter(f, fieldnames=results[0].keys())
            writer.writeheader()
            writer.writerows(results)

    print(f"\n生成ログ: {log_file}")

    # --- Cleanup: sanitize any existing output HTML files that still contain
    # Excel formula artifacts like =IFERROR(...) in <title> or meta description.
    try:
        # build id->brand_name map from results
        id_map = {r['id']: r['brand_name'] for r in results}

        out_dir = os.path.join('output', output_template)
        if os.path.isdir(out_dir):
            for fname in os.listdir(out_dir):
                if not fname.lower().endswith('.html'):
                    continue
                path = os.path.join(out_dir, fname)
                try:
                    with open(path, 'r', encoding='utf-8') as fh:
                        content = fh.read()

                    needs_write = False

                    # title
                    import re
                    title_match = re.search(r'<title>(.*?)</title>', content, re.IGNORECASE | re.DOTALL)
                    if title_match:
                        title_text = title_match.group(1)
                        if '=' in title_text or 'IFERROR' in title_text.upper():
                            # determine id from filename (assume first 5 chars)
                            file_id = fname[:5]
                            new_brand = id_map.get(file_id, '') or 'サロン'
                            new_title = f"{new_brand} - 本来のあなたへ"
                            content = content[:title_match.start(1)] + new_title + content[title_match.end(1):]
                            needs_write = True

                    # meta description
                    meta_match = re.search(r'<meta\s+name=["\']description["\']\s+content=["\'](.*?)["\']\s*/?>', content, re.IGNORECASE | re.DOTALL)
                    if meta_match:
                        meta_text = meta_match.group(1)
                        if '=' in meta_text or 'IFERROR' in meta_text.upper():
                            file_id = fname[:5]
                            new_brand = id_map.get(file_id, '') or 'サロン'
                            new_meta = f"{new_brand}の紹介ページです。"
                            content = content[:meta_match.start(1)] + new_meta + content[meta_match.end(1):]
                            needs_write = True

                    if needs_write:
                        with open(path, 'w', encoding='utf-8') as fh:
                            fh.write(content)
                except Exception:
                    # don't crash on unexpected files
                    pass
    except Exception:
        pass


if __name__ == '__main__':
    main()
